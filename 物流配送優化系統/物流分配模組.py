import streamlit as st
from datetime import datetime
from pymongo import MongoClient
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill

#MongoDB連線設定
client = MongoClient("mongodb://localhost:27017/")
db = client["my_database"]
drivers = db["drivers"]
proportion = db["numbox"]
routes = db["route"]
car = db["car"]

#初始session_state 
if "route_records" not in st.session_state:
    st.session_state["route_records"] = []
if "today_box" not in st.session_state:
    st.session_state.today_box = 0
if "editable_areas" not in st.session_state:
    st.session_state["editable_areas"] = []  # 當前可編輯地點清單

st.title("物流分配模組")
today = datetime.now().strftime("%Y/%m/%d")
st.markdown(f"今天日期：{today}")


input_today_box = st.number_input("輸入今日總箱數", key="input_today_box", value=st.session_state.today_box)
st.session_state.today_box = input_today_box
route_list = [r["路線"] for r in routes.find({}, {"_id": 0, "路線": 1})]
selected_route = st.selectbox("選擇路線", route_list, index=None, placeholder="請選擇路線")
numbox_list = [p["比例"] for p in proportion.find({}, {"_id": 0, "比例": 1})]

if selected_route:
    st.subheader(f"{selected_route} - 地點設定")

    if not st.session_state["editable_areas"]:
        route_data = routes.find_one({"路線": selected_route}, {"_id": 0, "地點": 1})
        st.session_state["editable_areas"] = route_data["地點"] if route_data and "地點" in route_data else []

    with st.expander("編輯地點列表", expanded=True):
        citys = [c["縣市"] for c in db["location"].find({}, {"_id": 0, "縣市": 1})]
        selected_city = st.selectbox("選擇縣市", citys, index=None, placeholder="請選擇縣市")
        area_list = []
        new_area = None
        if selected_city:
            city_data = db["location"].find_one({"縣市": selected_city}, {"_id": 0, "地點": 1})
            area_list = city_data["地點"] if city_data and "地點" in city_data else []
            new_area = st.selectbox("選擇要新增的地點", area_list, index=None, placeholder="請選擇地點")
        with st.form("add_area_form", clear_on_submit=True):
            col_add, col_reset = st.columns([1, 1])
            with col_add:
                add_clicked = st.form_submit_button("新增地點")
            with col_reset:
                reload_clicked = st.form_submit_button("重新載入原始路線")

            #新增地點到當前可編輯清單
            if add_clicked:
                if new_area and new_area not in st.session_state["editable_areas"]:
                    st.session_state["editable_areas"].append(new_area)
                    st.success(f"已新增地點：{new_area}")
                elif new_area in st.session_state["editable_areas"]:
                    st.warning("地點已存在於此路線中！")
                else:
                    st.warning("請先選擇地點！")

            #還原原始路線設定
            if reload_clicked:
                route_data = routes.find_one({"路線": selected_route}, {"_id": 0, "地點": 1})
                st.session_state["editable_areas"] = route_data["地點"] if route_data else []
                st.info("已還原為原始路線設定")

        #顯示目前地點，可刪除
        st.write("目前地點：")
        delete_area = None
        for area in st.session_state["editable_areas"]:
            col1, col2 = st.columns([4, 1])
            with col1:
                st.write(f"{area}")
            with col2:
                if st.button("刪除", key=f"del_{area}"):
                    delete_area = area

        if delete_area:
            st.session_state["editable_areas"].remove(delete_area)
            st.warning(f"已刪除地點：{delete_area}")
            st.rerun()

    #地點設定區 
    st.markdown("---")
    total_box = 0
    route_setting = {}

    for area in st.session_state["editable_areas"]:
        col1, col2, col3 = st.columns([2, 2, 2])
        with col1:
            st.write(f"{area}")
        with col2:
            selected_ratio = st.selectbox("比例", numbox_list, key=f"{selected_route}_{area}_ratio")
        with col3:
            extra_input = st.text_input("加貨", key=f"{selected_route}_{area}_extra", value="0")
        try:
            base_box = proportion.find_one({"比例": selected_ratio})["箱數"]
        except:
            base_box = 0
        try:
            extra_box = int(extra_input)
        except:
            extra_box = 0
        total = base_box + extra_box
        total_box += total
        route_setting[area] = {"比例": selected_ratio, "加貨": extra_box, "總箱數": total}

    st.markdown("---")
    st.subheader(f"路線總箱數：{total_box}")

    #指定駕駛人
    driver_list = [d["司機"] for d in drivers.find({}, {"_id": 0, "司機": 1})]
    selected_driver = st.selectbox("指定駕駛人", driver_list, index=None, placeholder="選擇駕駛")
    car_list = [d["車牌"] for d in car.find({}, {"_id": 0, "車牌": 1})]
    selected_car = st.selectbox("指定車輛", car_list, index=None, placeholder="選擇車輛")
    if selected_car:
        car_cap = car.find_one({"車牌": selected_car})["車容量"]
        if total_box > car_cap:
            st.error(f"超出車容量 {total_box - car_cap} 箱！ ({total_box}/{car_cap})")
        else:
            st.success(f"容量充足：{total_box}/{car_cap}")
            
    if st.button("新增到清單"):
        new_record = {
            "路線": selected_route,
            "駕駛人": selected_driver if selected_driver else "未指定",
            "車牌": selected_car if selected_car else "未指定",
            "設定": route_setting,
            "總箱數": total_box,
            "日期": today
        }
        st.session_state["route_records"].append(new_record)
        st.success(f"已儲存 {selected_route} 的設定！")

#可修改結算區
if st.session_state["route_records"]:
    st.markdown("---")
    st.subheader("結算紀錄（可修改）")
    
    total_all_box = 0
    delete_index = None
    
    for idx, rec in enumerate(st.session_state["route_records"]):
        temp_box = 0
        with st.expander(f"第 {idx+1} 筆：{rec['路線']}（{rec['駕駛人']}）"):
            st.write("地點與設定：")
            for area, info in rec["設定"].items():
                col1 , col2 ,col3 = st.columns([1,1,1])
                # 修改比例與加貨
                with col2:
                    rec["設定"][area]["比例"] = st.selectbox(
                    "選擇比例",
                    numbox_list,
                    index=numbox_list.index(info["比例"]) if info["比例"] in numbox_list else 0,
                    key=f"record_{idx}_{area}_ratio"
                    )
                with col3:
                    rec["設定"][area]["加貨"] = st.number_input(
                    "加貨",
                    min_value=0,
                    value=info["加貨"],
                    key=f"record_{idx}_{area}_extra"
                    )
                
                # 計算該地點總箱數
                base_box = proportion.find_one({"比例": rec["設定"][area]["比例"]})
                base_box = base_box["箱數"] if base_box else 0
                total_box_area = base_box + rec["設定"][area]["加貨"]
                rec["設定"][area]["總箱數"] = total_box_area
                temp_box += total_box_area
                
                with col1:
                    st.write(f"{area}：{total_box_area} 箱 \n\n(比例:{rec["設定"][area]["比例"]}為 {base_box}箱數 + 加貨 {rec['設定'][area]['加貨']})")
            
            total_all_box += temp_box
            
            car_cap = 0
            if rec["車牌"] != "未指定":
                car_data = car.find_one({"車牌": rec["車牌"]})
                car_cap = car_data["車容量"] if car_data else 0
            if temp_box <= car_cap:
                st.markdown(f'<span style="color:green; font-weight:bold">總箱數: {temp_box} / 車容量: {car_cap}</span>', unsafe_allow_html=True)
            else:
                st.markdown(f'<span style="color:red; font-weight:bold">總箱數: {temp_box} / 車容量: {car_cap} → 多 {temp_box - car_cap} 箱</span>', unsafe_allow_html=True)

            col_upd, col_del = st.columns([1, 1])
            with col_upd:
                if st.button(f"更新第 {idx+1} 筆", key=f"update_{idx}"):
                    st.success(f"已更新第 {idx+1} 筆設定")
            with col_del:
                if st.button(f"刪除第 {idx+1} 筆", key=f"delete_{idx}"):
                    delete_index = idx
    
    # 刪除紀錄
    if delete_index is not None:
        st.session_state["route_records"].pop(delete_index)
        st.warning(f"已刪除第 {delete_index+1} 筆紀錄")
        st.rerun()
    
    st.markdown("---")
    st.subheader("今日出貨統計")
    today_box = st.session_state.today_box
    if total_all_box > today_box:
        st.error(f"全日超出 {total_all_box - today_box} 箱！({total_all_box}/{today_box})")
    else:
        st.success(f"總箱數：{total_all_box}/{today_box}（未超出）")

#匯出Excel
if st.session_state["route_records"]:
    st.markdown("---")
    st.subheader("匯出所有結算結果")
    if st.button("匯出全部路線到 Excel"):
        export_data = []
        for rec in st.session_state["route_records"]:
            route_name = rec["路線"]
            driver_name = rec["駕駛人"]
            car_num = rec["車牌"]
            date = rec["日期"]
            car_cap = 0
            if car_num != "未指定":
                car_data = car.find_one({"車牌": car_num})
                car_cap = car_data["車容量"] if car_data else 0
            row = {
                "路線": route_name,
                "駕駛人": driver_name,
                "車牌": car_num,
                "路線總箱數": sum(info["總箱數"] for info in rec["設定"].values()),
                "車容量": car_cap
            }
            overload = row["路線總箱數"] - car_cap
            row["多出箱數"] = overload if overload > 0 else 0
            for idx, (area, info) in enumerate(rec["設定"].items(), start=1):
                base_box = proportion.find_one({"比例": info["比例"]})
                base_box = base_box["箱數"] if base_box else 0
                total_box_area = base_box + info["加貨"]
                row[f"地點{idx}"] = area
                row[f"比例{idx}"] = info["比例"]
                row[f"加貨{idx}"] = info["加貨"]
                row[f"地點總箱數{idx}"] = total_box_area
            export_data.append(row)
        df = pd.DataFrame(export_data)

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="出貨紀錄")
            workbook = writer.book
            sheet = workbook["出貨紀錄"]

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            total_col = None
            cap_col = None
            for idx, cell in enumerate(sheet[1], start=1):
                if cell.value == "路線總箱數":
                    total_col = idx
                elif cell.value == "車容量":
                    cap_col = idx

            if total_col and cap_col:
                for row_idx in range(2, sheet.max_row + 1):
                    total_value = sheet.cell(row=row_idx, column=total_col).value
                    cap_value = sheet.cell(row=row_idx, column=cap_col).value
                    if (
                        total_value
                        and cap_value
                        and isinstance(total_value, (int, float))
                        and isinstance(cap_value, (int, float))
                        and total_value > cap_value
                    ):
                        for col_idx in range(1, sheet.max_column + 1):
                            sheet.cell(row=row_idx, column=col_idx).fill = red_fill
        buffer.seek(0)
        st.download_button(
            label="下載 Excel（每條路線一列）",
            data=buffer,
            file_name=f"出貨紀錄_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
